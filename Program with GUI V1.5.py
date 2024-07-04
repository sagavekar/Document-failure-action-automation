# Author: Omkar Ashok Sagavekar - TSO
# Date: 2/10/2024

import xlwings as xw
from openpyxl import Workbook
from datetime import datetime as DT
import pandas as pd
import subprocess
from win32com.client import Dispatch as win32com_client_Dispatch
from os.path import abspath
from re import IGNORECASE, search
from sys import argv as sys_argv
from sys import exit as sys_exit
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QRadioButton,
    QFileDialog,
    QLabel,
    QVBoxLayout,
    QWidget,
    QMessageBox,
)


Assignee = {
    "AY": "Aditya Yadav",
    "OS": "Omkar Sagavekar",
    "MH": "Masroor Hafiz",
}

# Function to handle inbound sheet


def Inbound_auto(assignee_key, file_path):
    # Read the data
    df_inbound = pd.read_excel(file_path, sheet_name="Inbound Failures", skiprows=6)

    print("Inbound function execution started...")  # log

    # finding unique list of  DBname_LegalCompanyName_PartnerCode
    DBname_LegalCompanyName_PartnerCode = (
        df_inbound[
            (df_inbound["Assignee"] == Assignee.get(assignee_key))
            & (
                df_inbound["Responsibility"].str.lower() == "check with supplier"
            )  # to handle upper and lower case
            & (df_inbound["Status"] == "Pending")
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )

    total_lines = df_inbound[
        (df_inbound["Assignee"] == Assignee.get(assignee_key))
        & (
            df_inbound["Responsibility"].str.lower() == "check with supplier"
        )  # to handle upper and lower case
        & (df_inbound["Status"] == "Pending")
    ][["DBname_LegalCompanyName_PartnerCode"]][
        "DBname_LegalCompanyName_PartnerCode"
    ].count()

    total_files = len(DBname_LegalCompanyName_PartnerCode)
    print(f"Total files = {total_files}\nTotal lines = {total_lines}")
    # print(DBname_LegalCompanyName_PartnerCode)

    wb = xw.Book(file_path)
    Inbound_failures = wb.sheets["Inbound Failures"]

    """
    DBname_LegalCompanyName_PartnerCode = set()

    for row in Inbound_failures.used_range.rows[7:]:

        # wb1 = openpyxl.Workbook()
        # ws1 = wb1.create_sheet('Sheet1')

        if row[3].value == "Aditya Yadav" and row[5].value == "Check with Supplier" and row[7].value == "Pending":
            DBname_LegalCompanyName_PartnerCode.add(int(row[24].value))

    DBname_LegalCompanyName_PartnerCode = list(DBname_LegalCompanyName_PartnerCode)
    supplier_workbook = []

    """

    def email(P, target_wb_path):
        ol = win32com_client_Dispatch("outlook.application")  # from win32.client
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_Inbound document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(file_path, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1]).title()} team, <br><br>
            
        Please refer attached spreadsheet of Inbound document failures.<br>
        Error is mentioned in the file. Please make corrections and resend the documents.<br><br>
        
        <b>Customer name</b> : {P.split('_')[0]} <br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>


        Ignore this email if action is already taken. <br><br>
        
        Thanks and Regards, <br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        # newmail.Display()  # --> To display the mail
        newmail.Save()
        newmail.Close(0)  # 0 means close without sending
        print("Email draft created")
        # newmail.Send()

    # LOOP to create empty excel files
    for index, P in enumerate(DBname_LegalCompanyName_PartnerCode):
        # print(P)
        book = Workbook()  # book = openpyxl.Workbook()
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            "ObjectNumber",
            "ErrorMessage",
            "LegalCompanyName",
            "DocumentType",
            "SupplierInvoiceNumber",
            "Order Number",
            "DateCreated",
            "DBname_LegalCompanyName_PartnerCode",
            "OperationName",
            "StackTrace",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()
        print(f" {index+1} / {total_files} emty excel created")

    # loop to load data into excel file
    for index, P in enumerate(DBname_LegalCompanyName_PartnerCode):
        data_range = (
            Inbound_failures.range("A7:AF7")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee.get(assignee_key))  # Condition 1
            & (data_range["Responsibility"].str.lower() == "check with supplier")
            & (data_range["Status"] == "Pending")
            & (data_range["DBname_LegalCompanyName_PartnerCode"] == P)
        ]  # 4 condition applied

        selected_columns = filtered_data[
            [
                "DBname",
                "ObjectNumber",
                "ErrorMessage",
                "LegalCompanyName",
                "DocumentType",
                "SupplierInvoiceNumber",
                "Order Number",
                "DateCreated",
                "DBname_LegalCompanyName_PartnerCode",
                "OperationName",
                "StackTrace",
            ]
        ]

        # Open the target workbook
        target_wb = xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(index=False, header=True).value = selected_columns

        # Autofit the "DateCreated" column
        # Assuming "DateCreated" is in column I
        target_ws.range("H:H").autofit()

        # Disable text wrapping for the specified named range
        target_ws.range("A:K").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = abspath(target_wb.fullname)  # os.path.absoath
        target_wb.close()
        print(f"{index+1} / {total_files} excel data load completed")

        # calling email function
        email(P, target_wb_path)

    # Changing the status column from original file
    print("Main file data manupulation on progress... please wait....")
    counter_read = 0  # Counter to count against total lines
    for index, row in enumerate(Inbound_failures.used_range.rows[7:]):
        print("rows checked = ", {index + 1})
        if (
            row[3].value == Assignee.get(assignee_key)
            and str(row[5].value).lower() == "check with supplier"
            and row[7].value == "Pending"
        ):
            row[7].value = "No Document Found"  # Status column
            row[8].value = (
                "As per Resolution Column - Email sent to supplier to check"  # Action Performed column
            )
            # This will target the first action date
            row[11].value = DT.now().strftime("%m/%d/%Y")  # Action Date column
            counter_read += 1
            print(f"{counter_read} / {total_lines} completed")
        if counter_read == total_lines:  # --> Do not work because of enumerate
            break

    # wb.save() --> if main file needs to autosave
    # wb.close()  --> if main file needs to close

    print("Task completed ! Please check the drafted email and main excel file")

    # win32api.MessageBox(0, notification_text, notification_title, win32con.MB_OK)


# Function to handle outbound sheet
def Outbound_auto(assignee_key, file_path):
    # Read the data
    df_outbound = pd.read_excel(file_path, sheet_name="Outbond failures", skiprows=5)
    DBname_LegalCompanyName_PartnerCode = (
        df_outbound[
            (df_outbound["Assignee"] == Assignee.get(assignee_key))
            & (df_outbound["Status"] == "Pending")
            & (
                df_outbound["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
                | df_outbound["Error"].str.contains(
                    "but no response was received", case=False, regex=True
                )
                | df_outbound["Error"].str.contains(
                    "if they did not receive the PO", case=False, regex=True
                )
                | df_outbound["Error"].str.contains(
                    "if they received the PO", case=False, regex=True
                )
            )
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )
    # print(len(DBname_LegalCompanyName_PartnerCode),DBname_LegalCompanyName_PartnerCode)

    total_lines = df_outbound[
        (df_outbound["Assignee"] == Assignee.get(assignee_key))
        & (df_outbound["Status"] == "Pending")
        & (
            df_outbound["Error"].str.contains(
                "check with supplier", case=False, regex=True
            )
            | df_outbound["Error"].str.contains(
                "but no response was received", case=False, regex=True
            )
            | df_outbound["Error"].str.contains(
                "if they did not receive the PO", case=False, regex=True
            )
        )
    ][["DBname_LegalCompanyName_PartnerCode"]][
        "DBname_LegalCompanyName_PartnerCode"
    ].count()

    total_files = len(DBname_LegalCompanyName_PartnerCode)

    print(f"Total files = {total_files}\nTotal lines = {total_lines}")

    wb = xw.Book(file_path)
    Outbond_failures = wb.sheets["Outbond failures"]

    def email(P, target_wb_path, td_in_draft):
        ol = win32com_client_Dispatch("outlook.application")  # from win32com.client
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_PO document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(file_path, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1])} team, <br><br>
            
        Please refer attached spreadsheet of Purchase order failures.<br><br>
        We are not able to submit those POs via integration due to response error.<br>

        Kindly check and confirm if you have received POs already. 

        <br><br>
        
        <b>Customer name</b> : <u> {P.split('_')[0]} </u><br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>

        {td_in_draft.to_html(index=False)}

        <br>
        
        Thanks and Regards,<br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        # newmail.Display()  # --> To display the mail before sending it
        newmail.Save()
        newmail.Close(0)
        print("Email draft created")
        # newmail.Send()

    # loop to create emty excel files
    for index, P in enumerate(DBname_LegalCompanyName_PartnerCode):
        # print(P)
        book = Workbook()  # book = Workbook()  # this is from opepyxl
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            # "Error",
            "LegalCompanyName",
            "DocumentNumber",
            "OrderAmount",
            "DateModified",
            "PartnerCode",
            "DBname_LegalCompanyName_PartnerCode",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()
        print(f" {index+1} / {total_files} emty excel created")

    # loop to load data into excel files
    for index, P in enumerate(DBname_LegalCompanyName_PartnerCode):
        data_range = (
            Outbond_failures.range("A6:Y6")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee.get(assignee_key))  # Condition 1
            & (data_range["Status"] == "Pending")
            & (
                data_range["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
                | data_range["Error"].str.contains(
                    "but no response was received", case=False, regex=True
                )
                | data_range["Error"].str.contains(
                    "if they did not receive the PO", case=False, regex=True
                )
            )
            & (data_range["DBname_LegalCompanyName_PartnerCode"] == P)
        ]

        selected_columns = filtered_data[
            [
                "DBname",
                # "Error",
                "LegalCompanyName",
                "DocumentNumber",
                "OrderAmount",
                "DateModified",
                "PartnerCode",
                "DBname_LegalCompanyName_PartnerCode",
            ]
        ]

        td_in_draft = filtered_data[
            ["DBname","LegalCompanyName","DocumentNumber", "OrderAmount", "DateModified"]
        ]
        # Open the target workbook
        target_wb = xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(index=False, header=True).value = selected_columns

        # Autofit the "DateModified" column
        # Assuming "DateModified" is in column F
        target_ws.range("F:F").autofit()
        target_ws.range("B:B").autofit()  # Error column

        # Disable text wrapping for the specified named range
        target_ws.range("A:I").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = abspath(target_wb.fullname)  # os.path.abspath
        target_wb.close()
        print(f"{index+1} / {total_files} excel data load completed")

        # calling email function
        email(P, target_wb_path, td_in_draft)

    # Changing the status column from original file
    print("Main file data manupulation on progress... please wait....")
    counter_read = 0
    for index, row in enumerate(Outbond_failures.used_range.rows[6:]):
        print("rows checked = ", {index + 1})
        if (
            row[3].value == Assignee.get(assignee_key)
            # re.search and re.IGNORECASE
            and (
                search("check with supplier", str(row[4].value), IGNORECASE)
                or search(
                    "if they did not receive the PO", str(row[4].value), IGNORECASE
                )
                or search("but no response was received", str(row[4].value), IGNORECASE)
            )
            and row[5].value == "Pending"
        ):
            row[5].value = "In process"
            row[6].value = "Sending to supplier failed"
            row[7].value = "As per Resolution Column - Email sent to supplier to check"
            row[8].value = (
                f"{row[25].value}_PO document failure GEP/{(row[25].value).split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
            )
            row[9].value = DT.now().strftime("%m/%d/%Y")
            counter_read += 1
            print(f"{counter_read} / {total_lines} completed")
        if counter_read == total_lines:  #  --> do not work because of enumerate
            break

    print("Please wait for a momemt...all function have been executed")  # log


# ------Buidling the GUI starts here ----------------
class GUI(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Doc failure automation V1.5 | sagavekar.om@gmail.com")
        self.setGeometry(100, 100, 750, 550)

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()

        self.browse_button = QPushButton("Browse", self)
        self.browse_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;"
        )
        self.browse_button.clicked.connect(self.browse_file)
        self.layout.addWidget(self.browse_button)
        
        # trying to get current username using cmd "whoami" command 
        try:
            currentuser = str(subprocess.run("whoami", capture_output=True, text=True, shell=True).stdout).strip().split("\\")[1].split(".")[0]
            self.currentuser = QLabel(f'Welcome, {currentuser.capitalize() } !', self)
        except:
            self.currentuser = QLabel(f"Welcome !", self)
        self.currentuser.setStyleSheet("color: white;font-size: 21px;padding: 0px")
        self.currentuser.setWordWrap(True)
        self.layout.addWidget(self.currentuser)


        self.file_label = QLabel("Selected File: ", self)
        self.file_label.setStyleSheet("color: white;font-size: 23px;")
        self.file_label.setWordWrap(True)
        self.layout.addWidget(self.file_label)

        self.radio_buttons1 = []
        self.selected_radio_button = None  # To track selected radio button

        for key in Assignee:
            radio = QRadioButton(Assignee[key], self)
            radio.setStyleSheet("color: white;font-size: 23px;")
            radio.toggled.connect(
                lambda checked, button=radio, key=key: self.radio_button_selected(
                    key, button
                )
            )
            self.radio_buttons1.append(radio)
            self.layout.addWidget(radio)

        self.inbound_button = QPushButton("Inbound - check with supplier", self)
        self.inbound_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;"
        )
        self.inbound_button.clicked.connect(self.handle_inbound_auto)
        self.layout.addWidget(self.inbound_button)

        self.outbound_button = QPushButton("Outbound - check with supplier", self)
        self.outbound_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;"
        )
        self.outbound_button.clicked.connect(self.handle_outbound_auto)
        self.layout.addWidget(self.outbound_button)

        self.central_widget.setLayout(self.layout)
        self.central_widget.setStyleSheet("background-color: #222;")

    def browse_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("Excel files (*.xlsx)")
        # Restrict to existing files
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_path, _ = file_dialog.getOpenFileName()
        if file_path:
            self.file_label.setText("Selected File: " + file_path)

    def radio_button_selected(self, assignee_key, button):
        if button.isChecked():
            self.selected_radio_button = button
            self.selected_assignee_key = assignee_key
        else:
            self.selected_radio_button = None
            self.selected_assignee_key = None

    def handle_inbound_auto(self):
        # this to avoid calling below function if button is not select and file path / file lable in not selected
        if self.selected_radio_button and self.file_label.text() != "Selected File: ":

            assignee_key = next(
                key
                for key, value in Assignee.items()
                if value == self.selected_radio_button.text()
            )
            file_path = self.file_label.text().replace("Selected File: ", "")
            # from here program control will go to Inbound_auto function for instance.
            Inbound_auto(assignee_key, file_path)
            print("Inbound function execution completed")  # log
            QMessageBox.information(
                self, "Alert", "Inbound action completed ! Please check email draft"
            )

    def handle_outbound_auto(self):
        # this to avoid calling belo function if button is not select and file path / file lable in not selected
        if self.selected_radio_button and self.file_label.text() != "Selected File: ":
            assignee_key = next(
                key
                for key, value in Assignee.items()
                if value == self.selected_radio_button.text()
            )
            file_path = self.file_label.text().replace("Selected File: ", "")
            print("Outbound function has been called")  # log
            Outbound_auto(assignee_key, file_path)
            print("Outbound function execution completed")  # log
            QMessageBox.information(
                self, "Alert", "Outbound action completed ! Please check email draft"
            )


# ------Buidling the GUI ends here ----------------

if __name__ == "__main__":
    app = QApplication(sys_argv)
    window = GUI()
    window.show()
    sys_exit(app.exec_())
