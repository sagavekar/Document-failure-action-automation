import xlwings as xw
from openpyxl import Workbook
from datetime import datetime as DT
import pandas as pd
import win32com.client
import os
import re

# Define assingne here , user GUI/ input methods in advance version
Assignee = "Aditya Yadav"
main_file = "samples.xlsx"


def Inbound_auto():
    # Read the data
    df_inbound = pd.read_excel(main_file, sheet_name="Inbound Failures", skiprows=6)

    # finding unique list of  DBname_LegalCompanyName_PartnerCode
    DBname_LegalCompanyName_PartnerCode = (
        df_inbound[
            (df_inbound["Assignee"] == Assignee)
            & (df_inbound["Responsibility"] == "Check with Supplier")
            & (df_inbound["Status"] == "Pending")
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )
    # print(DBname_LegalCompanyName_PartnerCode)

    wb = xw.Book(main_file)
    Inbound_failures = wb.sheets["Inbound Failures"]

    """
    DBname_LegalCompanyName_PartnerCode = set()

    for row in Inbound_failures.used_range.rows[7:]:

        # wb1 = openpyxl.Workbook()
        # ws1 = wb1.create_sheet('Sheet1')

        if row[3].value == "Aditya Yadav" and row[5].value == "Check with Supplier" and row[7].value == "Pending":
            DBname_LegalCompanyName_PartnerCode.add(int(row[24].value))

    DBname_LegalCompanyName_PartnerCode = list(DBname_LegalCompanyName_PartnerCode)
    supplier_workbook = []

    """

    def email(P, target_wb_path):
        ol = win32com.client.Dispatch("outlook.application")
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_Inbound document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(main_file, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1]).title()} team, <br><br>
            
        Please refer attached speadsheet of Inbound document failures.<br>
        Error is mentioned in the file. Please make corrections and resend the documents.<br><br>
        
        <b>Customer name</b> : {P.split('_')[0]} <br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>


        Ignore this email if action is already taken. <br><br>
        
        Thanks and Regards, <br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        newmail.Display()  # --> To display the mail before sending it
        # newmail.Send()

    for P in DBname_LegalCompanyName_PartnerCode:
        # print(P)
        book = Workbook()
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            "ObjectNumber",
            "ErrorMessage",
            "LegalCompanyName",
            "DocumentType",
            "SupplierInvoiceNumber",
            "Order Number",
            "DateCreated",
            "DBname_LegalCompanyName_PartnerCode",
            "OperationName",
            "StackTrace",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()

    for P in DBname_LegalCompanyName_PartnerCode:
        data_range = (
            Inbound_failures.range("A7:AF7")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee)
            & (data_range["Responsibility"] == "Check with Supplier")  # Condition 1
            & (data_range["Status"] == "Pending")  # Condition 2
            & (  # Condition 3
                data_range["DBname_LegalCompanyName_PartnerCode"] == P
            )  # Condition 4
        ]

        selected_columns = filtered_data[
            [
                "DBname",
                "ObjectNumber",
                "ErrorMessage",
                "LegalCompanyName",
                "DocumentType",
                "SupplierInvoiceNumber",
                "Order Number",
                "DateCreated",
                "DBname_LegalCompanyName_PartnerCode",
                "OperationName",
                "StackTrace",
            ]
        ]

        # Open the target workbook
        target_wb = xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(index=False, header=True).value = selected_columns

        # Autofit the "DateCreated" column
        # Assuming "DateCreated" is in column I
        target_ws.range("H:H").autofit()

        # Disable text wrapping for the specified named range
        target_ws.range("A:K").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = os.path.abspath(target_wb.fullname)
        target_wb.close()

        # calling email function
        email(P, target_wb_path)

    # Changing the status column from original file
    for row in Inbound_failures.used_range.rows[7:]:
        if (
            row[3].value == Assignee
            and row[5].value == "Check with Supplier"
            and row[7].value == "Pending"
        ):
            row[7].value = "No Document Found"
            row[8].value = "As per Resolution Column - Email sent to supplier to check"
            row[11].value = DT.now().strftime("%m/%d/%Y")

    # wb.save() --> if main file needs to autosave
    # wb.close()  --> if main file needs to close

    print("Task completed ! Please check the drafted email and main excel file")

    # win32api.MessageBox(0, notification_text, notification_title, win32con.MB_OK)

def Outbound_auto():
    # Read the data
    df_outbound = pd.read_excel(main_file, sheet_name="Outbond failures", skiprows=5)
    DBname_LegalCompanyName_PartnerCode = (
        df_outbound[
            (df_outbound["Assignee"] == Assignee)
            & (df_outbound["Status"] == "Pending")
            & (
                df_outbound["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
            )
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )
    # print(len(DBname_LegalCompanyName_PartnerCode),DBname_LegalCompanyName_PartnerCode)

    wb = xw.Book(main_file)
    Outbond_failures = wb.sheets["Outbond failures"]


    def email(P, target_wb_path,td_in_draft):
        ol = win32com.client.Dispatch("outlook.application")
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_PO document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(main_file, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1])} team, <br><br>
            
        Please refer attached speadsheet of Purchase order failures.<br><br>
        We are not able to submit those via integration due to response error.<br>

        Details are mentioned in the file. Please check and confirm the status or amendments.<br><br>
        
        <b>Customer name</b> : <u> {P.split('_')[0]} </u><br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>

        {td_in_draft.to_html(index=False)}

        <br>
        
        Thanks and Regards,<br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        newmail.Display()  # --> To display the mail before sending it
        # newmail.Send()

    for P in DBname_LegalCompanyName_PartnerCode:
        # print(P)
        book = Workbook()
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            "Error",
            "LegalCompanyName",
            "DocumentNumber",
            "OrderAmount",
            "DateModified",
            "PartnerCode",
            "DBname_LegalCompanyName_PartnerCode",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()
        

    for P in DBname_LegalCompanyName_PartnerCode:
        data_range = (
            Outbond_failures.range("A6:X7")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee)  # Condition 1
            & (data_range["Status"] == "Pending")  # Condition 2
            & (
                data_range["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
            )  # Condition 3
            & (data_range["DBname_LegalCompanyName_PartnerCode"] == P)  # Condition 4
        ]

        selected_columns = filtered_data[
            [
                "DBname",
                "Error",
                "LegalCompanyName",
                "DocumentNumber",
                "OrderAmount",
                "DateModified",
                "PartnerCode",
                "DBname_LegalCompanyName_PartnerCode",
            ]
        ]

        td_in_draft = filtered_data[
            [
                "DocumentNumber",
                "OrderAmount",
                "Error",
                "DateModified"
            ]
        ]
        # Open the target workbook
        target_wb = xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(index=False, header=True).value = selected_columns

        # Autofit the "DateModified" column
        # Assuming "DateModified" is in column F
        target_ws.range("F:F").autofit()
        target_ws.range("B:B").autofit() # Error column

        # Disable text wrapping for the specified named range
        target_ws.range("A:I").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = os.path.abspath(target_wb.fullname)
        target_wb.close()

        # calling email function
        email(P, target_wb_path, td_in_draft)

    # Changing the status column from original file
    for row in Outbond_failures.used_range.rows[6:]:
        if (
            row[3].value == Assignee
            and re.search('check with supplier', str(row[4].value), re.IGNORECASE)
            and row[5].value == "Pending"
        ):
            row[5].value = "Sending to supplier failed"
            row[6].value = "As per Resolution Column - Email sent to supplier to check"
            row[7].value = DT.now().strftime("%m/%d/%Y")

Inbound_auto()
Outbound_auto()