import xlwings as xw
from openpyxl import Workbook
from datetime import datetime as DT
import pandas as pd
from win32com.client import Dispatch as win32com_client_Dispatch
from os.path import abspath
from re import IGNORECASE, search
from sys import argv as sys_argv
from sys import exit as sys_exit
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QRadioButton, QFileDialog, QLabel, QVBoxLayout, QWidget, QMessageBox, QProgressBar

Assignee = {
    "AY": "Aditya Yadav",
    "OS": "Omkar Sagavekar",
    "MH": "Masroor Hafiz",
}

# Function to handle inbound sheet


def Inbound_auto(assignee_key, file_path):
    # Read the data
    df_inbound = pd.read_excel(
        file_path, sheet_name="Inbound Failures", skiprows=6)

    # finding unique list of  DBname_LegalCompanyName_PartnerCode
    DBname_LegalCompanyName_PartnerCode = (
        df_inbound[
            (df_inbound["Assignee"] == Assignee.get(assignee_key))
            & (df_inbound["Responsibility"] == "Check with Supplier")
            & (df_inbound["Status"] == "Pending")
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )
    # print(DBname_LegalCompanyName_PartnerCode)

    wb = xw.Book(file_path)
    Inbound_failures = wb.sheets["Inbound Failures"]

    """
    DBname_LegalCompanyName_PartnerCode = set()

    for row in Inbound_failures.used_range.rows[7:]:

        # wb1 = openpyxl.Workbook()
        # ws1 = wb1.create_sheet('Sheet1')

        if row[3].value == "Aditya Yadav" and row[5].value == "Check with Supplier" and row[7].value == "Pending":
            DBname_LegalCompanyName_PartnerCode.add(int(row[24].value))

    DBname_LegalCompanyName_PartnerCode = list(DBname_LegalCompanyName_PartnerCode)
    supplier_workbook = []

    """

    def email(P, target_wb_path):
        ol = win32com_client_Dispatch(
            "outlook.application")  # from win32.client
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_Inbound document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(file_path, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["InboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1]).title()} team, <br><br>
            
        Please refer attached spreadsheet of Inbound document failures.<br>
        Error is mentioned in the file. Please make corrections and resend the documents.<br><br>
        
        <b>Customer name</b> : {P.split('_')[0]} <br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>


        Ignore this email if action is already taken. <br><br>
        
        Thanks and Regards, <br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        # newmail.Display()  # --> To display the mail
        newmail.Save()
        newmail.Close(0)  # 0 means close without sending
        # newmail.Send()

    for P in DBname_LegalCompanyName_PartnerCode:
        # print(P)
        book = Workbook()  # book = openpyxl.Workbook()
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            "ObjectNumber",
            "ErrorMessage",
            "LegalCompanyName",
            "DocumentType",
            "SupplierInvoiceNumber",
            "Order Number",
            "DateCreated",
            "DBname_LegalCompanyName_PartnerCode",
            "OperationName",
            "StackTrace",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()

    for P in DBname_LegalCompanyName_PartnerCode:
        data_range = (
            Inbound_failures.range("A7:AF7")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee.get(assignee_key))
            # Condition 1
            & (data_range["Responsibility"] == "Check with Supplier")
            & (data_range["Status"] == "Pending")  # Condition 2
            & (  # Condition 3
                data_range["DBname_LegalCompanyName_PartnerCode"] == P
            )  # Condition 4
        ]

        selected_columns = filtered_data[
            [
                "DBname",
                "ObjectNumber",
                "ErrorMessage",
                "LegalCompanyName",
                "DocumentType",
                "SupplierInvoiceNumber",
                "Order Number",
                "DateCreated",
                "DBname_LegalCompanyName_PartnerCode",
                "OperationName",
                "StackTrace",
            ]
        ]

        # Open the target workbook
        target_wb = xw.Book(f"{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(
            index=False, header=True).value = selected_columns

        # Autofit the "DateCreated" column
        # Assuming "DateCreated" is in column I
        target_ws.range("H:H").autofit()

        # Disable text wrapping for the specified named range
        target_ws.range("A:K").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = abspath(target_wb.fullname)  # os.path.absoath
        target_wb.close()

        # calling email function
        email(P, target_wb_path)

    # Changing the status column from original file
    for row in Inbound_failures.used_range.rows[7:]:
        if (
            row[3].value == Assignee.get(assignee_key)
            and row[5].value == "Check with Supplier"
            and row[7].value == "Pending"
        ):
            row[7].value = "No Document Found"
            row[8].value = "As per Resolution Column - Email sent to supplier to check"
            # This will target the first action date
            row[9].value = DT.now().strftime("%m/%d/%Y")

    # wb.save() --> if main file needs to autosave
    # wb.close()  --> if main file needs to close

    print("Task completed ! Please check the drafted email and main excel file")

    # win32api.MessageBox(0, notification_text, notification_title, win32con.MB_OK)

# Function to handle outbound sheet


def Outbound_auto(assignee_key, file_path):
    # Read the data
    df_outbound = pd.read_excel(
        file_path, sheet_name="Outbond failures", skiprows=5)
    DBname_LegalCompanyName_PartnerCode = (
        df_outbound[
            (df_outbound["Assignee"] == Assignee.get(assignee_key))
            & (df_outbound["Status"] == "Pending")
            & (
                df_outbound["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
            )
        ][["DBname_LegalCompanyName_PartnerCode"]][
            "DBname_LegalCompanyName_PartnerCode"
        ]
        .unique()
        .tolist()
    )
    # print(len(DBname_LegalCompanyName_PartnerCode),DBname_LegalCompanyName_PartnerCode)

    wb = xw.Book(file_path)
    Outbond_failures = wb.sheets["Outbond failures"]

    def email(P, target_wb_path, td_in_draft):
        ol = win32com_client_Dispatch(
            "outlook.application")  # from win32com.client
        olmailitem = 0x0  # size of the new email
        newmail = ol.CreateItem(olmailitem)
        newmail.Subject = f"{P}_PO document failure GEP/{P.split('_')[0]}| Reporting_date {DT.now().strftime('%m.%d.%Y')}"
        try:
            df_contact = pd.read_excel(file_path, sheet_name="Contact")

            newmail.To = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_To"]].to_string(index=False, header=False)

            newmail.CC = df_contact[
                (df_contact["DBname_LegalCompanyName_PartnerCode"] == P)
            ][["OutboundEmail_CC"]].to_string(index=False, header=False)
        except:
            newmail.To = "Add email address"
            newmail.CC = "Add email address"

        # Email draft in HTML formate
        newmail.HTMLBody = f"""
        <html>
        Hi {str(P.split('_')[1])} team, <br><br>
            
        Please refer attached spreadsheet of Purchase order failures.<br><br>
        We are not able to submit those via integration due to response error.<br>

        Details are mentioned in the file. Please check and confirm the status or amendments.<br><br>
        
        <b>Customer name</b> : <u> {P.split('_')[0]} </u><br>
        <b>Vendor/supplier name</b> : {P.split('_')[1]} <br>
        <b>Integration type</b> : cXML/EDI  <br>
        <b>Reporting date</b>: {DT.now().strftime('%m.%d.%Y')} <br><br>

        {td_in_draft.to_html(index=False)}

        <br>
        
        Thanks and Regards,<br>
        Supplier Enablement Team_GEP 
        </html>
        """
        attach = target_wb_path
        newmail.Attachments.Add(attach)
        # newmail.Display()  # --> To display the mail before sending it
        newmail.Save()
        newmail.Close(0)
        # newmail.Send()

    for P in DBname_LegalCompanyName_PartnerCode:
        # print(P)
        book = Workbook()  # book = Workbook()  # this is from opepyxl
        # sheet = book.create_sheet(f'{P}')
        # supplier_worksheet.append(sheet)

        headers = [
            "DBname",
            "Error",
            "LegalCompanyName",
            "DocumentNumber",
            "OrderAmount",
            "DateModified",
            "PartnerCode",
            "DBname_LegalCompanyName_PartnerCode",
        ]
        book.active.append(headers)

        # supplier_workbook.append(book)
        book.save(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).columns.autofit()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").sheets["Sheet"].range(
            "A1:K1"
        ).api.Font.Bold = True
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").save()
        xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx").close()

    for P in DBname_LegalCompanyName_PartnerCode:
        data_range = (
            Outbond_failures.range("A6:Y6")
            .expand()
            .options(pd.DataFrame, index=False, header=True)
            .value
        )

        filtered_data = data_range[
            (data_range["Assignee"] == Assignee.get(
                assignee_key))  # Condition 1
            & (data_range["Status"] == "Pending")  # Condition 2
            & (
                data_range["Error"].str.contains(
                    "check with supplier", case=False, regex=True
                )
            )  # Condition 3
            # Condition 4
            & (data_range["DBname_LegalCompanyName_PartnerCode"] == P)
        ]

        selected_columns = filtered_data[
            [
                "DBname",
                "Error",
                "LegalCompanyName",
                "DocumentNumber",
                "OrderAmount",
                "DateModified",
                "PartnerCode",
                "DBname_LegalCompanyName_PartnerCode",
            ]
        ]

        td_in_draft = filtered_data[
            [
                "DocumentNumber",
                "OrderAmount",
                "Error",
                "DateModified"
            ]
        ]
        # Open the target workbook
        target_wb = xw.Book(f"PO_{P}_{DT.now().strftime('%m.%d.%Y')}.xlsx")
        target_ws = target_wb.sheets.active

        # Write the filtered data from the DataFrame to the target worksheet
        target_ws.range("A1").options(
            index=False, header=True).value = selected_columns

        # Autofit the "DateModified" column
        # Assuming "DateModified" is in column F
        target_ws.range("F:F").autofit()
        target_ws.range("B:B").autofit()  # Error column

        # Disable text wrapping for the specified named range
        target_ws.range("A:I").api.WrapText = False

        # Save and close the target workbook
        target_wb.save()
        target_wb_path = abspath(target_wb.fullname)  # os.path.abspath
        target_wb.close()

        # calling email function
        email(P, target_wb_path, td_in_draft)

    # Changing the status column from original file
    for row in Outbond_failures.used_range.rows[6:]:
        if (
            row[3].value == Assignee.get(assignee_key)
            # re.search and re.IGNORECASE
            and search('check with supplier', str(row[4].value), IGNORECASE)
            and row[5].value == "Pending"
        ):
            row[5].value = "In process"
            row[6].value = "Sending to supplier failed"
            row[7].value = "As per Resolution Column - Email sent to supplier to check"
            row[8].value = DT.now().strftime("%m/%d/%Y")


# ------Buidling the GUI ----------------
class GUI(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Doc failure automation | Omkar.sagavekar@gep.com")
        self.setGeometry(100, 100, 750, 550)

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()

        self.browse_button = QPushButton("Browse", self)
        self.browse_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;")
        self.browse_button.clicked.connect(self.browse_file)
        self.layout.addWidget(self.browse_button)

        self.file_label = QLabel("Selected File: ", self)
        self.file_label.setStyleSheet("color: white;font-size: 23px;")
        self.file_label.setWordWrap(True)
        self.layout.addWidget(self.file_label)

        self.radio_buttons1 = []
        self.selected_radio_button = None  # To track selected radio button

        for key in Assignee:
            radio = QRadioButton(Assignee[key], self)
            radio.setStyleSheet("color: white;font-size: 23px;")
            radio.toggled.connect(
                lambda checked, button=radio, key=key: self.radio_button_selected(key, button))
            self.radio_buttons1.append(radio)
            self.layout.addWidget(radio)

        self.inbound_button = QPushButton(
            "Inbound - check with supplier", self)
        self.inbound_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;")
        self.inbound_button.clicked.connect(self.handle_inbound_auto)
        self.layout.addWidget(self.inbound_button)

        self.outbound_button = QPushButton(
            "Outbound - check with supplier", self)
        self.outbound_button.setStyleSheet(
            "background-color: #333; color: white; padding: 10px;font-size: 23px;")
        self.outbound_button.clicked.connect(self.handle_outbound_auto)
        self.layout.addWidget(self.outbound_button)

        self.central_widget.setLayout(self.layout)
        self.central_widget.setStyleSheet("background-color: #222;")

    def browse_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("Excel files (*.xlsx)")
        # Restrict to existing files
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_path, _ = file_dialog.getOpenFileName()
        if file_path:
            self.file_label.setText("Selected File: " + file_path)

    def radio_button_selected(self, assignee_key, button):
        if button.isChecked():
            self.selected_radio_button = button
            self.selected_assignee_key = assignee_key
        else:
            self.selected_radio_button = None
            self.selected_assignee_key = None

    def handle_inbound_auto(self):
        # this to avoid calling belo function if button is not select and file path / file lable in not selected
        if self.selected_radio_button and self.file_label.text() != "Selected File: ":

            assignee_key = next(key for key, value in Assignee.items(
            ) if value == self.selected_radio_button.text())
            file_path = self.file_label.text().replace("Selected File: ", "")
            Inbound_auto(assignee_key, file_path)
            QMessageBox.information(
                self, "Alert", "Inbound action completed ! Please check email draft")

    def handle_outbound_auto(self):
        # this to avoid calling belo function if button is not select and file path / file lable in not selected
        if self.selected_radio_button and self.file_label.text() != "Selected File: ":
            assignee_key = next(key for key, value in Assignee.items(
            ) if value == self.selected_radio_button.text())
            file_path = self.file_label.text().replace("Selected File: ", "")
            Outbound_auto(assignee_key, file_path)
            QMessageBox.information(
                self, "Alert", "Outbound action completed ! Please check email draft")


if __name__ == "__main__":
    app = QApplication(sys_argv)
    window = GUI()
    window.show()
    sys_exit(app.exec_())
